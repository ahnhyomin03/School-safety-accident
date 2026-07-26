from collections import Counter, defaultdict
import json
import openpyxl
import re
import xml.etree.ElementTree as ET
import zipfile


PUBLIC_XLSX = "doc/data/2024_스포츠안전사고_실태조사_체육인_raw.xlsx"
OUTPUT_JSON = "doc/data/2024_스포츠안전사고_공개자료_축구분석.json"


def pct(n, d):
    return round(n * 100 / d, 2) if d else None


def excel_column_number(cell_reference):
    letters = re.match(r"[A-Z]+", cell_reference).group(0)
    number = 0
    for letter in letters:
        number = number * 26 + ord(letter) - 64
    return number


def cell_value(cell, namespace, shared_strings):
    value_element = cell.find(f"{namespace}v")
    if value_element is None:
        inline = cell.find(f"{namespace}is")
        return (
            "".join(text.text or "" for text in inline.iter(f"{namespace}t"))
            if inline is not None
            else None
        )
    raw = value_element.text
    if cell.attrib.get("t") == "s":
        return shared_strings[int(raw)]
    if raw is None:
        return None
    try:
        numeric = float(raw)
        return int(numeric) if numeric.is_integer() else numeric
    except ValueError:
        return raw


def iter_selected_xlsx_rows(
    path, selected_columns, early_filter_column=None, early_filter_value=None
):
    namespace = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
    with zipfile.ZipFile(path) as archive:
        shared_strings = []
        if "xl/sharedStrings.xml" in archive.namelist():
            with archive.open("xl/sharedStrings.xml") as handle:
                for _, element in ET.iterparse(handle, events=("end",)):
                    if element.tag == f"{namespace}si":
                        shared_strings.append(
                            "".join(text.text or "" for text in element.iter(f"{namespace}t"))
                        )
                        element.clear()

        with archive.open("xl/worksheets/sheet1.xml") as handle:
            for _, element in ET.iterparse(handle, events=("end",)):
                if element.tag != f"{namespace}row":
                    continue
                row_number = int(element.attrib["r"])
                if row_number <= 4:
                    element.clear()
                    continue

                cells = element.findall(f"{namespace}c")
                if early_filter_column is not None:
                    filter_value = None
                    for cell in cells[:10]:
                        if excel_column_number(cell.attrib["r"]) == early_filter_column:
                            filter_value = cell_value(cell, namespace, shared_strings)
                            break
                    if filter_value != early_filter_value:
                        element.clear()
                        continue

                result = {}
                for cell in cells:
                    column = excel_column_number(cell.attrib["r"])
                    if column not in selected_columns:
                        continue
                    result[column] = cell_value(cell, namespace, shared_strings)
                yield result
                element.clear()


def public_sports_summary():
    wb = openpyxl.load_workbook(PUBLIC_XLSX, read_only=True, data_only=True)

    guide = wb["GUIDE"]
    body_items = {}
    for row in guide.iter_rows(values_only=True):
        variable = row[1]
        if isinstance(variable, str) and variable.startswith("Q7#") and variable != "Q7#ETC_38":
            body_items[variable] = {
                "code": row[4],
                "label": str(row[5]).strip(),
            }

    sheet = wb["DATA"]
    rows = sheet.iter_rows(max_row=4, values_only=True)
    header_rows = list(rows)
    variables = header_rows[3]
    index = {value: i for i, value in enumerate(variables) if value is not None}

    required = [
        "SQ1", "SQ2_R", "SQ4", "Q6", "Q6_1", "Q13__51", "Q17_1",
        "Q25", "Q26", "Q27", "Q27_1#1", "Q27_1#2", "Q27_1#3",
        "Q28", "Q35", "Q38",
    ] + list(body_items)
    missing = [name for name in required if name not in index]
    if missing:
        raise RuntimeError(f"Missing variables: {missing}")
    selected_columns = {index[name] + 1 for name in required}

    groups = {
        "all_soccer": {"age_codes": set(range(1, 9))},
        "child_12_or_younger_soccer": {"age_codes": {1}},
        "teen_13_18_soccer": {"age_codes": {2}},
        "teen_13_18_lifestyle_soccer": {"age_codes": {2}, "athlete_codes": {1}},
        "teen_13_18_professional_soccer": {"age_codes": {2}, "athlete_codes": {2}},
        "school_age_18_or_younger_soccer": {"age_codes": {1, 2}},
        "adult_19_64_soccer": {"age_codes": {3, 4, 5, 6, 7}},
        "adult_19_64_lifestyle_soccer": {
            "age_codes": {3, 4, 5, 6, 7},
            "athlete_codes": {1},
        },
        "adult_19_64_professional_soccer": {
            "age_codes": {3, 4, 5, 6, 7},
            "athlete_codes": {2},
        },
        "senior_65_plus_soccer": {"age_codes": {8}},
    }
    counts = {
        name: {
            "respondents": 0,
            "injured": 0,
            "body": Counter(),
            "warmup_before_injury": Counter(),
            "usual_warmup": Counter(),
            "protective_equipment": Counter(),
            "environment_check": Counter(),
            "manual_check": Counter(),
            "leader_present": Counter(),
            "athlete_type": Counter(),
            "surface": Counter(),
            "knee_by_warmup": defaultdict(Counter),
            "warmup_quality": defaultdict(list),
        }
        for name in groups
    }

    for sparse_row in iter_selected_xlsx_rows(
        PUBLIC_XLSX,
        selected_columns,
        early_filter_column=index["SQ2_R"] + 1,
        early_filter_value=51,
    ):
        def value(variable):
            return sparse_row.get(index[variable] + 1)

        age_code = value("SQ4")
        for group_name, spec in groups.items():
            if age_code not in spec["age_codes"]:
                continue
            if "athlete_codes" in spec and value("SQ1") not in spec["athlete_codes"]:
                continue
            bucket = counts[group_name]
            bucket["respondents"] += 1
            bucket["usual_warmup"][value("Q38")] += 1
            bucket["athlete_type"][value("SQ1")] += 1
            if value("Q6") != 1:
                continue

            bucket["injured"] += 1
            selected_parts = []
            for variable, item in body_items.items():
                label = item["label"]
                if value(variable) == item["code"]:
                    bucket["body"][label] += 1
                    selected_parts.append(label)

            warmup = value("Q27")
            knee = "무릎" in selected_parts
            bucket["warmup_before_injury"][warmup] += 1
            bucket["protective_equipment"][value("Q25")] += 1
            bucket["environment_check"][value("Q26")] += 1
            bucket["manual_check"][value("Q28")] += 1
            bucket["leader_present"][value("Q17_1")] += 1
            bucket["surface"][value("Q13__51")] += 1
            bucket["knee_by_warmup"][warmup]["knee" if knee else "other"] += 1

            for variable in ("Q27_1#1", "Q27_1#2", "Q27_1#3"):
                score = value(variable)
                if isinstance(score, (int, float)):
                    bucket["warmup_quality"][variable].append(score)

    result = {}
    for group_name, bucket in counts.items():
        injured = bucket["injured"]
        body_mentions = sum(bucket["body"].values())
        body_table = []
        for label, n in bucket["body"].most_common():
            body_table.append(
                {
                    "body_part": label,
                    "count": n,
                    "injured_person_pct": pct(n, injured),
                    "body_mention_share_pct": pct(n, body_mentions),
                }
            )
        result[group_name] = {
            "respondents": bucket["respondents"],
            "injured": injured,
            "injury_experience_pct": pct(injured, bucket["respondents"]),
            "body_mentions": body_mentions,
            "body_parts": body_table,
            "warmup_before_injury": dict(bucket["warmup_before_injury"]),
            "usual_warmup": dict(bucket["usual_warmup"]),
            "protective_equipment": dict(bucket["protective_equipment"]),
            "environment_check": dict(bucket["environment_check"]),
            "manual_check": dict(bucket["manual_check"]),
            "leader_present": dict(bucket["leader_present"]),
            "athlete_type": dict(bucket["athlete_type"]),
            "surface": dict(bucket["surface"]),
            "knee_by_warmup": {
                str(key): dict(value)
                for key, value in bucket["knee_by_warmup"].items()
            },
            "warmup_quality_mean": {
                variable: round(sum(values) / len(values), 3) if values else None
                for variable, values in bucket["warmup_quality"].items()
            },
        }
    return result


if __name__ == "__main__":
    output = {"public_2024_sports_survey": public_sports_summary()}
    with open(OUTPUT_JSON, "w", encoding="utf-8") as handle:
        json.dump(output, handle, ensure_ascii=False, indent=2)
    print(OUTPUT_JSON)
